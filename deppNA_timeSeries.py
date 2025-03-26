import openpyxl as xl
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense
import tensorflow as tf
from sklearn.preprocessing import MinMaxScaler
import pandas as pd
import numpy as np

def deepNA(file):
    wb = xl.load_workbook(file)
    print("Sheets available: ", wb.sheetnames)
    sheet_name = input('Please enter selected sheet name: ')
    ws = wb[sheet_name]
    
    columns = [ws.cell(row=1, column=i).value.strip() if ws.cell(row=1, column=i).value else "" for i in range(1, ws.max_column + 1)]
    print("Columns available: ", columns)
    
    col = input("Please enter selected column to fill in: ").strip() 

    try:
        col_idx = columns.index(col) + 1
    except ValueError:
        print(f"Column '{col}' not found in the sheet.")
        return

    data = [ws.cell(row=i, column=col_idx).value for i in range(2, ws.max_row + 1)]
    print(f"First ten values: {data[:10]}")

    data = [None if value is None or value == "" else value for value in data]

    df = pd.DataFrame(data, columns=[col])

    scaler = MinMaxScaler()
    df_scaled = scaler.fit_transform(df.dropna().values.reshape(-1, 1))

    def create_sequences(data, seq_length=10):
        X, y = [], []
        for i in range(len(data) - seq_length):
            X.append(data[i:i + seq_length])
            y.append(data[i + seq_length])
        return np.array(X), np.array(y)

    seq_length = 10
    X, y = create_sequences(df_scaled)

    model = Sequential([LSTM(50, return_sequences=True, input_shape=(seq_length, 1)),
                        LSTM(50, return_sequences=False),
                        Dense(25, activation="relu"),
                        Dense(1)])
    model.compile(optimizer="adam", loss="mse")

    model.fit(X, y, epochs=20, batch_size=16)

    missing = df[df[col].isna()].index

    if len(missing) > 0 and missing[-1] >= len(df) - 50:
        mean_value = df[col].dropna().iloc[:10].mean()  
        df.loc[missing, col] = mean_value
        print(f"Last {len(missing)} values were filled with the mean: {mean_value}")
    else:
        for i in missing:
            if i >= seq_length:
                input_seq = df_scaled[i - seq_length:i]
                if input_seq.shape[0] == seq_length:
                    input_seq = input_seq.reshape(1, seq_length, 1)
                    pred = model.predict(input_seq)
                    df.at[i, col] = scaler.inverse_transform(pred)[0][0]
                else:
                    print(f"Skipping index {i} due to insufficient data points for sequence.")
            else:
                print(f"Skipping index {i} as there are not enough previous data points to form a sequence.")

    for i, value in enumerate(df[col], start=2):
        ws.cell(row=i, column=col_idx, value=value)

    wb.save("Filled_" + file)
    print(f"File successfully saved as 'Filled_{file}'")

